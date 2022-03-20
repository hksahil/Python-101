#!/usr/bin/env python
# coding: utf-8

# ## Working with Excel files

# - If you just want to manipulate excel's data,use pd.read_excel() and convert it to dataframe and do the manipulation
# - openpyxl is the best module to work with excel's other things,like changing sheet names,formatting,automating it etc
# 
# - indexes start at 1 not 0

# In[1]:


get_ipython().system('pip install openpyxl')


# In[2]:


import openpyxl


# In[3]:


import os


# ### Changing the working directory

# In[4]:


os.chdir('C:\\users\\Sahil Choudhary\\Desktop')


# ### Open Excel file

# In[5]:


# Use load_workbook function and it will give the workbook object
wb=openpyxl.load_workbook('example.xlsx')


# In[6]:


print(type(wb))


# ### Get sheet names

# In[7]:


wb.sheetnames


# ### Create Sheet

# In[8]:


wb.create_sheet(title='My new Sheet',index=1)


# ### Get one of the Sheet

# In[9]:


sheet=wb['Sheet1']


# In[10]:


print(type(sheet))


# ### Changing sheet name

# In[11]:


sheet.title='new name'


# ### Get the number of rows and columns

# In[12]:


print(sheet.max_row)
print(sheet.max_column)


# ### Getting Values

# In[13]:


# Getting cell object
sheet['B1']


# In[14]:


print(sheet['B1'].value)
print(sheet.cell(row=1,column=2).value)


# In[15]:


# Getting its value
sheet['A1'].value


# ### Changing Values

# In[16]:


sheet['B1'].value='Sahil Choudhary'
# By defauly,saves it in internal memory


# ### Save the result as excel file

# In[17]:


wb.save('example2.xlsx')


# ### Looping through cells

# In[18]:


# Get the first 5 cells of 2nd column
for i in range(1,5):
    print(sheet.cell(row=i,column=2).value)


# In[19]:


ws=wb['Sheet2']


# In[20]:


rows=ws.iter_rows(min_row=1,max_row=7,min_col=1,max_col=2)
# we have iter_cols as well
# returns generator object


# In[21]:


for row in rows:
    print(row)
    # returns tupple of cell locations


# In[22]:


for a,b in rows:
    print(a.value,b.value)
    # returns tupple of cell locations


# ### Get column name by number

# In[23]:


print(openpyxl.utils.cell.get_column_letter(1))


# ### Change format of cells

# In[24]:


# create font object and give it to cell
from openpyxl.styles import Font
sheet['B1'].font=Font(sz=14,bold=True,italic=True)

# save it as new file in end


# ### Adding Border

# In[25]:


from openpyxl.styles import Border,Side

# Pass color and border style to Side
# top=Side(border_style='thin|thick|medium|dashed|double',color='hexcode')
top=Side(border_style='thin')

# Border(top=top,bottom=bottom,left=left,right=right)
border=Border(top=top)

sheet['B1'].border=border

